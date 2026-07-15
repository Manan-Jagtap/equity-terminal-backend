"""
app/export_routes.py — Excel workbook exports (openpyxl).

  GET /api/export/screener.xlsx  → one "Screener" sheet, one row per company
                                   with a precomputed Valuation + MarketSnapshot.
  GET /api/export/{ticker}.xlsx  → full company workbook: a formatted Summary,
                                   a LIVE self-justifying DCF/RI model (every
                                   number is an in-cell formula off an editable
                                   assumptions block — change a driver and the
                                   intrinsic recomputes), and the P&L / Balance
                                   Sheet / Cash Flow with derived margin & growth
                                   rows computed by formula.

Design goals for the company workbook:
  * Numbers are numbers (never strings), with real Excel number formats.
  * The valuation is not a static dump — it is a working model. The forecast
    schedule, terminal value and the equity bridge are all `=` formulas that
    reference the assumptions, so the sheet TELLS and JUSTIFIES the DCF, and at
    the engine's own inputs it reproduces the engine's intrinsic exactly.
  * Clean, light, print-ready styling: teal section bands, amber input cells,
    hidden gridlines, frozen headers — aesthetically in line with the one-pager.
"""
import io

from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app import models, engines
from app import sector_params as SP
from app.assemble import build_company, effective_assumptions, _shape_statements
from app.consensus import analyst_consensus

router = APIRouter(prefix="/api/export", tags=["export"])

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ── palette (matches the one-pager) ──────────────────────────────────────────
TEAL    = "0F766E"
TEAL_LT = "E3F0EE"
INK     = "141719"
MUTE    = "6B7280"
GREY_LT = "F4F6F7"
LINE_C  = "D9DEE3"
INPUT   = "FFF7E6"     # amber wash → "this is an editable input"
GREEN   = "15803D"
RED     = "B91C1C"
WHITE   = "FFFFFF"

_thin = Side(style="thin", color=LINE_C)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

F_TITLE  = Font(bold=True, size=15, color=WHITE)
F_SUB    = Font(size=9, color="D8ECE9")
F_SEC    = Font(bold=True, size=10, color=WHITE)
F_LBL    = Font(size=10, color=INK)
F_MUTE   = Font(size=9, color=MUTE)
F_VAL    = Font(size=10, color=INK)
F_INPUT  = Font(size=10, color="92400E")
F_BOLD   = Font(bold=True, size=10, color=INK)
F_HEAD   = Font(bold=True, size=9, color=WHITE)
F_RESULT = Font(bold=True, size=13, color=TEAL)
F_NOTE   = Font(italic=True, size=8, color=MUTE)

# number formats
CR   = '#,##0.0'
CR0  = '#,##0'
PS   = '#,##0.00'
PCT  = '0.0%'
PCT2 = '0.00%'
NUMX = '0.00"x"'
INTF = '0'


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _num(x, nd=4):
    if x is None:
        return None
    try:
        return round(float(x), nd)
    except (TypeError, ValueError):
        return None


def _theme(ws, col_widths):
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _w(ws, coord, value=None, font=None, fmt=None, fill_c=None,
       align=None, border=False):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fmt:
        c.number_format = fmt
    if fill_c:
        c.fill = _fill(fill_c)
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center")
    if border:
        c.border = BORDER
    return c


def _band(ws, row, text, last_col="H", font=F_SEC, fill_c=TEAL, h=18):
    """Full-width coloured section band across A:last_col."""
    ws.merge_cells(f"A{row}:{last_col}{row}")
    _w(ws, f"A{row}", text, font=font, fill_c=fill_c,
       align="left")
    ws.row_dimensions[row].height = h
    # paint the merged range so the fill spans
    for col in range(1, ws[f"{last_col}1"].column + 1):
        ws.cell(row=row, column=col).fill = _fill(fill_c)


def _input(ws, row, label, value, fmt, note=None):
    """A labelled editable input: amber value cell in column B."""
    _w(ws, f"A{row}", label, font=F_LBL, align="left")
    _w(ws, f"B{row}", value, font=F_INPUT, fmt=fmt, fill_c=INPUT,
       align="right", border=True)
    if note:
        _w(ws, f"C{row}", note, font=F_NOTE, align="left")
    return f"B{row}"


def _derived(ws, row, label, formula, fmt, font=F_VAL, result=False, note=None):
    """A labelled formula/derived cell (teal-tinted for results)."""
    _w(ws, f"A{row}", label, font=(F_BOLD if result else F_LBL), align="left")
    c = _w(ws, f"B{row}", formula, font=(F_RESULT if result else font),
           fmt=fmt, align="right", border=True,
           fill_c=(TEAL_LT if result else None))
    if note:
        _w(ws, f"C{row}", note, font=F_NOTE, align="left")
    return f"B{row}"


# ── Screener export ──────────────────────────────────────────────────────────

_SCREENER_HEADER = ["Ticker", "Name", "Sector", "Valuation Sector", "Price",
                    "Intrinsic", "MoS %", "Verdict", "Composite", "P/E", "P/B",
                    "ROE %", "Analyst Target", "Analyst Rating"]


@router.get("/screener.xlsx")
def export_screener(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Screener"
    ws.append(_SCREENER_HEADER)
    widths = [14, 30, 22, 18, 12, 12, 10, 12, 11, 9, 9, 9, 14, 14]
    _theme(ws, widths)
    for cell in ws[1]:
        cell.font = F_HEAD
        cell.fill = _fill(TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    price_by = {m.company_id: m.price for m in db.query(models.MarketSnapshot).all()}
    val_by = {}
    try:
        val_by = {v.company_id: v for v in db.query(models.Valuation).all()}
    except Exception:
        db.rollback()

    for co in db.query(models.Company).order_by(models.Company.ticker).all():
        v = val_by.get(co.id)
        price = price_by.get(co.id)
        if v is None or price is None:
            continue
        ws.append([
            co.ticker, co.name, co.sector, v.valuation_sector,
            _num(price, 2), _num(v.intrinsic, 2),
            _num(v.mos * 100 if v.mos is not None else None, 1),
            v.verdict, _num(v.composite, 1),
            _num(v.pe, 2), _num(v.pb, 2),
            _num(v.roe * 100 if v.roe is not None else None, 1),
            _num(v.analyst_target, 2), v.analyst_rating,
        ])
    # number formats on the data columns
    for row in ws.iter_rows(min_row=2):
        for j, fmt in ((5, PS), (6, PS), (7, '0.0'), (9, '0.0'),
                       (10, NUMX), (11, NUMX), (12, '0.0'), (13, PS)):
            if row[j - 1].value is not None:
                row[j - 1].number_format = fmt
    return _xlsx_response(wb, "screener.xlsx")


# ── Company export: Summary ──────────────────────────────────────────────────

def _summary_sheet(ws, co, price, rec, analyst, fair_value):
    """`fair_value` is either a cross-sheet formula string (e.g. "='DCF Model'!B50")
    linking live to the model, or a plain number when no model sheet was built."""
    _theme(ws, [30, 22, 26, 14, 14, 14, 14, 14])
    v = (rec or {}).get("valuation") or {}
    linked = isinstance(fair_value, str) and fair_value.startswith("=")
    verdict = (rec or {}).get("verdict") or "—"

    _w(ws, "A1", co.name, font=F_TITLE, fill_c=TEAL, align="left")
    ws.merge_cells("A1:H1")
    for col in range(1, 9):
        ws.cell(row=1, column=col).fill = _fill(TEAL)
    ws.row_dimensions[1].height = 26
    _w(ws, "A2", f"{co.ticker}  ·  {co.sector}", font=F_SUB, fill_c=TEAL, align="left")
    ws.merge_cells("A2:H2")
    for col in range(1, 9):
        ws.cell(row=2, column=col).fill = _fill(TEAL)

    # Headline fair value = the app's blended estimate (matches the one-pager
    # and the product UI). The pure-DCF model value is shown alongside, linked
    # live to the model sheet, so the two never silently disagree.
    blended = _num((rec or {}).get("intrinsic"), 2)
    r = 4
    _band(ws, r, "VERDICT & FAIR VALUE"); r += 1
    rows = [
        ("Verdict", verdict, None, F_BOLD, None),
        ("Current price (Rs)", _num(price, 2), PS, F_VAL, None),
        ("Fair value / share (Rs)", blended, PS, F_RESULT, None),
        ("DCF model value (Rs)", fair_value, PS, F_VAL,
         "computed live on the model sheet" if linked else None),
        ("Margin of safety",
         _num((rec or {}).get("mos"), 4) if (rec or {}).get("mos") is not None else None,
         PCT, F_VAL, None),
        ("Composite score", _num((rec or {}).get("composite"), 1), '0.0', F_VAL, None),
        ("Primary method", (rec or {}).get("primary_method") or v.get("method"), None, F_VAL, None),
        ("Valuation sector", (rec or {}).get("valuation_sector"), None, F_VAL, None),
    ]
    for lbl, val, fmt, font, note in rows:
        _w(ws, f"A{r}", lbl, font=F_LBL, align="left")
        _w(ws, f"B{r}", val, font=font, fmt=fmt, align="right", border=True,
           fill_c=(TEAL_LT if font is F_RESULT else None))
        if note:
            _w(ws, f"C{r}", note, font=F_NOTE, align="left")
        r += 1

    r += 1
    _band(ws, r, "ANALYST CONSENSUS"); r += 1
    a = analyst or {}
    for lbl, val, fmt in [
        ("Target (Rs)", _num(a.get("target"), 2), PS),
        ("Range low / high (Rs)",
         (f"{_num(a.get('low'),0)} – {_num(a.get('high'),0)}"
          if a.get("low") is not None else None), None),
        ("Rating", a.get("rating"), None),
        ("Implied upside", _num(a.get("upside"), 4) if a.get("upside") is not None else None, PCT),
        ("# Analysts", _num(a.get("n"), 0), INTF),
    ]:
        _w(ws, f"A{r}", lbl, font=F_LBL, align="left")
        _w(ws, f"B{r}", val, font=F_VAL, fmt=fmt, align="right", border=True)
        r += 1

    _w(ws, f"A{r + 1}",
       ("Fair value blends the DCF model with relative multiples and analyst "
        "consensus; the DCF model value is computed live on the model sheet. "
        if linked else
        "Fair value is the engine's blended estimate. ") +
       "Educational use only; not SEBI-registered investment advice.",
       font=F_NOTE, align="left")


# ── Company export: the live valuation model ─────────────────────────────────

def _fcff_model_sheet(ws, ticker, co, a, v, price, mature_roic):
    """A working two-stage FCFF DCF — every figure is an in-cell formula off the
    amber assumptions, so it recomputes when a driver is edited and reproduces
    the engine's intrinsic at the engine's inputs."""
    _theme(ws, [30, 15, 14, 14, 14, 14, 13, 15])
    N = max(3, round(a["fade_years"]))

    _w(ws, "A1", f"{ticker} — FCFF Discounted Cash Flow", font=F_TITLE,
       fill_c=TEAL, align="left")
    ws.merge_cells("A1:H1")
    for col in range(1, 9):
        ws.cell(row=1, column=col).fill = _fill(TEAL)
    ws.row_dimensions[1].height = 24

    r = 3
    _band(ws, r, "ASSUMPTIONS  (amber cells are editable — the model re-runs)"); r += 1
    R = {}
    R["rev0"]  = _input(ws, r, "Base revenue (Rs Cr)", _num(co["revenue"], 2), CR); r += 1
    R["g1"]    = _input(ws, r, "Stage-1 revenue growth", _num(a["rev_growth"], 6), PCT); r += 1
    R["gt"]    = _input(ws, r, "Terminal growth", _num(a["terminal_growth"], 6), PCT); r += 1
    R["ebitm"] = _input(ws, r, "EBIT margin", _num(a["ebit_margin"], 6), PCT); r += 1
    R["tax"]   = _input(ws, r, "Tax rate", _num(a["tax_rate"], 6), PCT); r += 1
    R["reinv"] = _input(ws, r, "Reinvestment rate (stage 1)", _num(a["reinvest_rate"], 6), PCT); r += 1
    R["N"]     = _input(ws, r, "Forecast horizon N (yrs)", N, INTF); r += 1
    R["N1"]    = _derived(ws, r, "Franchise phase N1 (yrs)", f"=MAX(1,INT({R['N']}/2))", INTF,
                          note="high-growth years before the fade"); r += 1
    R["mroic"] = _input(ws, r, "Mature ROIC (terminal)", _num(mature_roic, 6), PCT); r += 1
    R["trr"]   = _derived(ws, r, "Terminal reinvestment",
                          f"=IFERROR(MIN(MAX({R['gt']}/{R['mroic']},0),0.75),{R['reinv']})", PCT,
                          note="g / mature ROIC, capped at 0.75"); r += 1
    R["nd"]    = _input(ws, r, "Net debt (Rs Cr)", _num(co["net_debt"], 2), CR); r += 1
    R["sh"]    = _input(ws, r, "Shares outstanding (Cr)", _num(co["shares"], 4), CR); r += 1
    R["px"]    = _input(ws, r, "Current price (Rs)", _num(price, 2), PS); r += 1

    r += 1
    _band(ws, r, "COST OF CAPITAL"); r += 1
    R["rf"]   = _input(ws, r, "Risk-free (10Y G-sec)", _num(a["risk_free"], 6), PCT); r += 1
    R["beta"] = _input(ws, r, "Beta", _num(a["beta"], 4), '0.00'); r += 1
    R["erp"]  = _input(ws, r, "Equity risk premium", _num(a["erp"], 6), PCT); r += 1
    R["ke"]   = _derived(ws, r, "Cost of equity  Ke", f"={R['rf']}+{R['beta']}*{R['erp']}", PCT,
                         note="Rf + β · ERP"); r += 1
    R["dw"]   = _input(ws, r, "Debt weight", _num(a["debt_weight"], 6), PCT); r += 1
    R["cod"]  = _input(ws, r, "Cost of debt (pre-tax)", _num(a["cost_debt"], 6), PCT); r += 1
    R["wacc"] = _derived(ws, r, "WACC",
                         f"=(1-{R['dw']})*{R['ke']}+{R['dw']}*{R['cod']}*(1-{R['tax']})", PCT,
                         font=F_BOLD, note="(1-Wd)·Ke + Wd·Kd·(1-t)"); r += 1

    # ── forecast & discounting table ─────────────────────────────────────────
    r += 1
    _band(ws, r, "FORECAST & DISCOUNTING  (Rs Cr)"); r += 1
    ht = r
    headers = ["Year", "Rev growth", "Revenue", "EBIT", "NOPAT", "FCFF",
               "Disc factor", "PV of FCFF"]
    for j, h in enumerate(headers, start=1):
        _w(ws, f"{get_column_letter(j)}{ht}", h, font=F_HEAD, fill_c=MUTE,
           align="center", border=True)
    ws.row_dimensions[ht].height = 15
    first = ht + 1
    for i in range(1, N + 1):
        rr = ht + i
        _w(ws, f"A{rr}", i, font=F_VAL, fmt=INTF, align="center", border=True)
        _w(ws, f"B{rr}",
           f"=IFERROR(IF(A{rr}<={R['N1']},{R['g1']},"
           f"{R['g1']}+({R['gt']}-{R['g1']})*((A{rr}-{R['N1']})/({R['N']}-{R['N1']}))),{R['gt']})",
           font=F_VAL, fmt=PCT, align="right", border=True)
        prev = R["rev0"] if i == 1 else f"C{rr-1}"
        _w(ws, f"C{rr}", f"={prev}*(1+B{rr})", font=F_VAL, fmt=CR, align="right", border=True)
        _w(ws, f"D{rr}", f"=C{rr}*{R['ebitm']}", font=F_VAL, fmt=CR, align="right", border=True)
        _w(ws, f"E{rr}", f"=D{rr}*(1-{R['tax']})", font=F_VAL, fmt=CR, align="right", border=True)
        _w(ws, f"F{rr}", f"=E{rr}*(1-{R['reinv']})", font=F_VAL, fmt=CR, align="right", border=True)
        _w(ws, f"G{rr}", f"=IF(A{rr}<={R['N']},1/(1+{R['wacc']})^A{rr},0)",
           font=F_MUTE, fmt='0.000', align="right", border=True)
        _w(ws, f"H{rr}", f"=F{rr}*G{rr}", font=F_BOLD, fmt=CR, align="right", border=True)
    last = ht + N
    nopat_rng = f"E{first}:E{last}"

    # ── terminal value ───────────────────────────────────────────────────────
    r = last + 2
    _band(ws, r, "TERMINAL VALUE  (Rs Cr)"); r += 1
    R["tfcff"] = _derived(ws, r, "Terminal-year FCFF",
                          f"=INDEX({nopat_rng},{R['N']})*(1+{R['gt']})*(1-{R['trr']})", CR,
                          note="NOPAT(N) · (1+g) · (1 − terminal reinvest)"); r += 1
    R["tv"]    = _derived(ws, r, "Terminal value (undiscounted)",
                          f"=IF({R['wacc']}>{R['gt']},{R['tfcff']}/({R['wacc']}-{R['gt']}),0)", CR,
                          note="Gordon growth: FCFF / (WACC − g)"); r += 1
    R["tvpv"]  = _derived(ws, r, "PV of terminal value",
                          f"={R['tv']}/(1+{R['wacc']})^{R['N']}", CR); r += 1

    # ── equity bridge ────────────────────────────────────────────────────────
    r += 1
    _band(ws, r, "VALUATION BRIDGE"); r += 1
    R["pvsum"] = _derived(ws, r, "PV of explicit FCFF", f"=SUM(H{first}:H{last})", CR); r += 1
    R["ev"]    = _derived(ws, r, "Enterprise value", f"={R['pvsum']}+{R['tvpv']}", CR,
                          font=F_BOLD); r += 1
    _w(ws, f"A{r}", "less: Net debt", font=F_LBL, align="left")
    _w(ws, f"B{r}", f"=-{R['nd']}", font=F_VAL, fmt=CR, align="right", border=True); r += 1
    R["eq"]    = _derived(ws, r, "Equity value", f"={R['ev']}-{R['nd']}", CR, font=F_BOLD); r += 1
    _derived(ws, r, "÷ Shares outstanding (Cr)", f"={R['sh']}", CR); r += 1
    R["iv"]    = _derived(ws, r, "INTRINSIC VALUE / SHARE (Rs)",
                          f"=IFERROR({R['eq']}/{R['sh']},0)", PS, result=True); r += 1
    _derived(ws, r, "Current price (Rs)", f"={R['px']}", PS); r += 1
    R["mos"]   = _derived(ws, r, "Margin of safety",
                          f"=IFERROR({R['iv']}/{R['px']}-1,0)", PCT, font=F_BOLD); r += 1

    # reconciliation with the engine
    r += 1
    _w(ws, f"A{r}", "Engine intrinsic (for reconciliation):", font=F_MUTE, align="left")
    _w(ws, f"B{r}", _num(v.get("intrinsic"), 2), font=F_MUTE, fmt=PS, align="right"); r += 1
    _w(ws, f"A{r}",
       "This sheet recomputes the DCF live from the assumptions above; at the "
       "engine's inputs cell B (intrinsic) matches the engine value.",
       font=F_NOTE, align="left")
    ws.merge_cells(f"A{r}:H{r}")
    return R["iv"]


def _ri_model_sheet(ws, ticker, co, a, v, price):
    """A working two-stage Residual-Income model for financials — same live,
    formula-driven design as the FCFF sheet."""
    _theme(ws, [32, 15, 15, 15, 13, 15, 15, 12])
    N = max(3, round(a["fade_years"]))

    _w(ws, "A1", f"{ticker} — Residual Income (excess-return) model",
       font=F_TITLE, fill_c=TEAL, align="left")
    ws.merge_cells("A1:H1")
    for col in range(1, 9):
        ws.cell(row=1, column=col).fill = _fill(TEAL)
    ws.row_dimensions[1].height = 24

    r = 3
    _band(ws, r, "ASSUMPTIONS  (amber cells are editable — the model re-runs)"); r += 1
    R = {}
    R["eq0"]  = _input(ws, r, "Shareholder equity (Rs Cr)", _num(co["equity"], 2), CR); r += 1
    R["sh"]   = _input(ws, r, "Shares outstanding (Cr)", _num(co["shares"], 4), CR); r += 1
    R["bvps"] = _derived(ws, r, "Book value / share (Rs)", f"={R['eq0']}/{R['sh']}", PS); r += 1
    R["froe"] = _input(ws, r, "Forecast ROE (franchise)", _num(a["forecast_roe"], 6), PCT); r += 1
    R["troe"] = _input(ws, r, "Terminal ROE", _num(a["terminal_roe"], 6), PCT); r += 1
    R["pay"]  = _input(ws, r, "Dividend payout", _num(a["payout"], 6), PCT); r += 1
    R["ret"]  = _derived(ws, r, "Retention (1 − payout)", f"=1-{R['pay']}", PCT); r += 1
    R["gt"]   = _input(ws, r, "Terminal growth", _num(a["terminal_growth"], 6), PCT); r += 1
    R["N"]    = _input(ws, r, "Forecast horizon N (yrs)", N, INTF); r += 1
    R["N1"]   = _derived(ws, r, "Franchise phase N1 (yrs)", f"=MAX(1,INT({R['N']}/2))", INTF); r += 1
    R["px"]   = _input(ws, r, "Current price (Rs)", _num(price, 2), PS); r += 1

    r += 1
    _band(ws, r, "COST OF EQUITY"); r += 1
    R["rf"]   = _input(ws, r, "Risk-free (10Y G-sec)", _num(a["risk_free"], 6), PCT); r += 1
    R["beta"] = _input(ws, r, "Beta", _num(a["beta"], 4), '0.00'); r += 1
    R["erp"]  = _input(ws, r, "Equity risk premium", _num(a["erp"], 6), PCT); r += 1
    R["ke"]   = _derived(ws, r, "Cost of equity  Ke", f"={R['rf']}+{R['beta']}*{R['erp']}", PCT,
                         font=F_BOLD, note="Rf + β · ERP"); r += 1

    r += 1
    _band(ws, r, "EXCESS-RETURN SCHEDULE  (per share, Rs)"); r += 1
    ht = r
    headers = ["Year", "ROE", "BV begin", "Residual income", "Disc factor",
               "PV of RI", "BV end", ""]
    for j, h in enumerate(headers, start=1):
        if h:
            _w(ws, f"{get_column_letter(j)}{ht}", h, font=F_HEAD, fill_c=MUTE,
               align="center", border=True)
    ws.row_dimensions[ht].height = 15
    first = ht + 1
    for i in range(1, N + 1):
        rr = ht + i
        _w(ws, f"A{rr}", i, font=F_VAL, fmt=INTF, align="center", border=True)
        _w(ws, f"B{rr}",
           f"=IFERROR(IF(A{rr}<={R['N1']},{R['froe']},"
           f"{R['froe']}+({R['troe']}-{R['froe']})*((A{rr}-{R['N1']})/({R['N']}-{R['N1']}))),{R['troe']})",
           font=F_VAL, fmt=PCT, align="right", border=True)
        bvb = R["bvps"] if i == 1 else f"G{rr-1}"
        _w(ws, f"C{rr}", f"={bvb}", font=F_VAL, fmt=PS, align="right", border=True)
        _w(ws, f"D{rr}", f"=(B{rr}-{R['ke']})*C{rr}", font=F_VAL, fmt=PS, align="right", border=True)
        _w(ws, f"E{rr}", f"=IF(A{rr}<={R['N']},1/(1+{R['ke']})^A{rr},0)",
           font=F_MUTE, fmt='0.000', align="right", border=True)
        _w(ws, f"F{rr}", f"=D{rr}*E{rr}", font=F_BOLD, fmt=PS, align="right", border=True)
        _w(ws, f"G{rr}", f"=C{rr}*(1+B{rr}*{R['ret']})", font=F_VAL, fmt=PS, align="right", border=True)
    last = ht + N
    bvend_rng = f"G{first}:G{last}"

    r = last + 2
    _band(ws, r, "TERMINAL VALUE  (per share, Rs)"); r += 1
    R["rin"]  = _derived(ws, r, "Terminal residual income",
                         f"=({R['troe']}-{R['ke']})*INDEX({bvend_rng},{R['N']})", PS,
                         note="(terminal ROE − Ke) · BV(N)"); r += 1
    R["tv"]   = _derived(ws, r, "Terminal value (undiscounted)",
                         f"=IF({R['ke']}>{R['gt']},{R['rin']}/({R['ke']}-{R['gt']}),0)", PS); r += 1
    R["tvpv"] = _derived(ws, r, "PV of terminal value",
                         f"={R['tv']}/(1+{R['ke']})^{R['N']}", PS); r += 1

    r += 1
    _band(ws, r, "VALUATION BRIDGE  (per share, Rs)"); r += 1
    _derived(ws, r, "Book value / share (t0)", f"={R['bvps']}", PS); r += 1
    R["pvsum"] = _derived(ws, r, "PV of explicit residual income", f"=SUM(F{first}:F{last})", PS); r += 1
    _derived(ws, r, "PV of terminal value", f"={R['tvpv']}", PS); r += 1
    R["iv"] = _derived(ws, r, "INTRINSIC VALUE / SHARE (Rs)",
                       f"={R['bvps']}+{R['pvsum']}+{R['tvpv']}", PS, result=True); r += 1
    _derived(ws, r, "Current price (Rs)", f"={R['px']}", PS); r += 1
    _derived(ws, r, "Margin of safety", f"=IFERROR({R['iv']}/{R['px']}-1,0)", PCT, font=F_BOLD); r += 1

    r += 1
    _w(ws, f"A{r}", "Engine intrinsic (for reconciliation):", font=F_MUTE, align="left")
    _w(ws, f"B{r}", _num(v.get("intrinsic"), 2), font=F_MUTE, fmt=PS, align="right")
    return R["iv"]


# ── Company export: statements in standard order, with derived rows ──────────
# Canonical top-to-bottom ordering (raw line-item keys). Anything not listed is
# appended after, alphabetically — the statements never render in random order.
_PL_ORDER = ["revenue", "interest_income", "gross_profit", "other_income", "nii",
             "sga", "provisions", "ebitda", "depreciation", "ebit",
             "interest_expense", "pbt", "tax", "pat", "net_profit"]
_BS_ORDER = ["cash", "cash_equivalents", "short_term_investments", "investments",
             "receivables", "inventory", "prepaid_expenses", "other_current_assets",
             "current_assets", "ppe_gross", "accumulated_depreciation",
             "fixed_assets", "goodwill", "intangibles", "total_assets",
             "payables", "notes_payable_st", "accrued_expenses", "current_lt_debt",
             "other_current_liabilities", "current_liabilities", "lt_debt",
             "borrowings", "total_debt", "total_liabilities",
             "reserves", "net_worth", "equity", "total_equity"]
_CF_ORDER = ["operating_cf", "capex", "investing_cf", "financing_cf", "fcf",
             "net_change_cash"]
_ORDER = {"PL": _PL_ORDER, "BS": _BS_ORDER, "CF": _CF_ORDER}

_BOLD_KEYS = {"revenue", "gross_profit", "nii", "ebitda", "ebit", "pbt", "pat",
              "net_profit", "current_assets", "total_assets", "current_liabilities",
              "total_liabilities", "net_worth", "equity", "total_equity",
              "operating_cf", "fcf", "net_change_cash"}

# balance-sheet grouping → sub-band before the first row of each group
_BS_GROUP = {}
for _k in _BS_ORDER[:_BS_ORDER.index("total_assets") + 1]:
    _BS_GROUP[_k] = "ASSETS"
for _k in _BS_ORDER[_BS_ORDER.index("payables"):_BS_ORDER.index("total_liabilities") + 1]:
    _BS_GROUP[_k] = "LIABILITIES"
for _k in _BS_ORDER[_BS_ORDER.index("reserves"):]:
    _BS_GROUP[_k] = "EQUITY"

_LABELS = {
    "revenue": "Revenue", "interest_income": "Interest income",
    "gross_profit": "Gross profit", "other_income": "Other income",
    "nii": "Net interest income", "sga": "SG&A / operating expenses",
    "provisions": "Provisions", "ebitda": "EBITDA",
    "depreciation": "Depreciation & amortisation", "ebit": "EBIT",
    "interest_expense": "Interest expense", "pbt": "Profit before tax",
    "tax": "Tax", "pat": "PAT (net profit)", "net_profit": "Net profit",
    "cash": "Cash & bank", "cash_equivalents": "Cash equivalents",
    "short_term_investments": "Short-term investments", "investments": "Investments",
    "receivables": "Receivables", "inventory": "Inventory",
    "prepaid_expenses": "Prepaid expenses", "other_current_assets": "Other current assets",
    "current_assets": "Total current assets", "ppe_gross": "PP&E (gross)",
    "accumulated_depreciation": "Less: accumulated depreciation",
    "fixed_assets": "Net fixed assets", "goodwill": "Goodwill",
    "intangibles": "Intangibles", "total_assets": "TOTAL ASSETS",
    "payables": "Payables", "notes_payable_st": "Short-term notes payable",
    "accrued_expenses": "Accrued expenses", "current_lt_debt": "Current portion of LT debt",
    "other_current_liabilities": "Other current liabilities",
    "current_liabilities": "Total current liabilities", "lt_debt": "Long-term debt",
    "borrowings": "Borrowings", "total_debt": "Total debt",
    "total_liabilities": "TOTAL LIABILITIES", "reserves": "Reserves & surplus",
    "net_worth": "Net worth", "equity": "Shareholders' equity",
    "total_equity": "Total equity", "operating_cf": "Cash from operations",
    "capex": "Capital expenditure", "investing_cf": "Cash from investing",
    "financing_cf": "Cash from financing", "fcf": "Free cash flow",
    "net_change_cash": "Net change in cash",
}


def _label(key):
    return _LABELS.get(key, str(key).replace("_", " ").capitalize())


def _ordered_items(keys, stmt_type):
    order = _ORDER.get(stmt_type, [])
    idx = {k: i for i, k in enumerate(order)}
    known = sorted([k for k in keys if k in idx], key=lambda k: idx[k])
    unknown = sorted([k for k in keys if k not in idx])
    return known + unknown


def _statement_sheet(ws, statements, stmt_type, add_derived=False):
    years = sorted(statements.keys())
    if not years:
        _w(ws, "A1", "No statement data available.", font=F_MUTE)
        return {}
    ncol = len(years) + 1
    last_col = get_column_letter(ncol)
    _theme(ws, [34] + [13] * len(years))

    _w(ws, "A1", "Line Item (Rs Cr)", font=F_HEAD, fill_c=TEAL, align="left", border=True)
    for j, y in enumerate(years, start=2):
        _w(ws, f"{get_column_letter(j)}1", f"FY{str(y)[-2:]}", font=F_HEAD,
           fill_c=TEAL, align="center", border=True)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 16

    keys = {item for y in years for item in (statements[y].get(stmt_type) or {})}
    items = _ordered_items(keys, stmt_type)
    row_of = {}
    r = 2
    group_seen = set()
    for key in items:
        # balance-sheet group sub-band (ASSETS / LIABILITIES / EQUITY)
        if stmt_type == "BS":
            grp = _BS_GROUP.get(key)
            if grp and grp not in group_seen:
                group_seen.add(grp)
                _band(ws, r, grp, last_col=last_col, font=F_HEAD, fill_c=MUTE, h=14)
                r += 1
        bold = key in _BOLD_KEYS
        _w(ws, f"A{r}", _label(key), font=(F_BOLD if bold else F_LBL), align="left")
        for j, y in enumerate(years, start=2):
            val = (statements[y].get(stmt_type) or {}).get(key)
            _w(ws, f"{get_column_letter(j)}{r}", _num(val, 2),
               font=(F_BOLD if bold else F_VAL), fmt=CR, align="right", border=True,
               fill_c=(GREY_LT if bold else None))
        row_of[key] = r
        r += 1

    if not add_derived:
        return row_of

    rev = row_of.get("revenue")
    ebitda = row_of.get("ebitda")
    ebit = row_of.get("ebit")
    pat = row_of.get("pat") or row_of.get("net_profit")
    if not rev:
        return
    r += 1
    _band(ws, r, "PROFITABILITY & GROWTH  (computed by formula)", last_col=last_col); r += 1
    for lbl, num_r in (("EBITDA margin", ebitda), ("EBIT margin", ebit), ("PAT margin", pat)):
        if not num_r:
            continue
        _w(ws, f"A{r}", lbl, font=F_LBL, align="left")
        for j in range(2, ncol + 1):
            cl = get_column_letter(j)
            _w(ws, f"{cl}{r}", f'=IFERROR({cl}{num_r}/{cl}{rev},"")',
               font=F_VAL, fmt=PCT, align="right", border=True)
        r += 1
    for lbl, base_r in (("Revenue growth YoY", rev), ("PAT growth YoY", pat)):
        if not base_r:
            continue
        _w(ws, f"A{r}", lbl, font=F_LBL, align="left")
        for j in range(3, ncol + 1):
            cl, pv = get_column_letter(j), get_column_letter(j - 1)
            _w(ws, f"{cl}{r}", f'=IFERROR({cl}{base_r}/{pv}{base_r}-1,"")',
               font=F_VAL, fmt=PCT, align="right", border=True)
        r += 1
    return row_of


# ── Company export: a Ratios sheet, computed live off the statement sheets ───

def _ratios_sheet(ws, years, pl_rows, bs_rows, pl_title, bs_title):
    """Cross-sheet ratio grid: every cell is a formula off the P&L / Balance
    Sheet tabs, so the ratios stay tied to the statements they come from."""
    if not years:
        _w(ws, "A1", "No data for ratios.", font=F_MUTE)
        return
    ncol = len(years) + 1
    last_col = get_column_letter(ncol)
    _theme(ws, [30] + [13] * len(years))

    _w(ws, "A1", "Ratio", font=F_HEAD, fill_c=TEAL, align="left", border=True)
    for j, y in enumerate(years, start=2):
        _w(ws, f"{get_column_letter(j)}1", f"FY{str(y)[-2:]}", font=F_HEAD,
           fill_c=TEAL, align="center", border=True)
    ws.freeze_panes = "B2"

    P = lambda k, j: f"'{pl_title}'!{get_column_letter(j)}{pl_rows[k]}" if pl_rows.get(k) else None
    B = lambda k, j: f"'{bs_title}'!{get_column_letter(j)}{bs_rows[k]}" if bs_rows.get(k) else None
    equity_key = "net_worth" if bs_rows.get("net_worth") else "equity"
    debt_key = "total_debt" if bs_rows.get("total_debt") else "borrowings"

    # (label, section, fmt, builder(j) → formula-or-None)
    rows = [
        ("PROFITABILITY", None, None, None),
        ("Gross margin", "band", PCT, lambda j: (f"={P('gross_profit',j)}/{P('revenue',j)}" if P('gross_profit',j) and P('revenue',j) else None)),
        ("EBITDA margin", None, PCT, lambda j: (f"={P('ebitda',j)}/{P('revenue',j)}" if P('ebitda',j) and P('revenue',j) else None)),
        ("EBIT margin", None, PCT, lambda j: (f"={P('ebit',j)}/{P('revenue',j)}" if P('ebit',j) and P('revenue',j) else None)),
        ("PAT margin", None, PCT, lambda j: (f"={P('pat',j)}/{P('revenue',j)}" if P('pat',j) and P('revenue',j) else None)),
        ("RETURNS", None, None, None),
        ("Return on equity", "band", PCT, lambda j: (f"={P('pat',j)}/{B(equity_key,j)}" if P('pat',j) and B(equity_key,j) else None)),
        ("Return on assets", None, PCT, lambda j: (f"={P('pat',j)}/{B('total_assets',j)}" if P('pat',j) and B('total_assets',j) else None)),
        ("Return on capital employed", None, PCT, lambda j: (f"={P('ebit',j)}/({B(equity_key,j)}+{B(debt_key,j)})" if P('ebit',j) and B(equity_key,j) and B(debt_key,j) else None)),
        ("Asset turnover", None, NUMX, lambda j: (f"={P('revenue',j)}/{B('total_assets',j)}" if P('revenue',j) and B('total_assets',j) else None)),
        ("LEVERAGE & COVERAGE", None, None, None),
        ("Debt / equity", "band", NUMX, lambda j: (f"={B(debt_key,j)}/{B(equity_key,j)}" if B(debt_key,j) and B(equity_key,j) else None)),
        ("Interest coverage", None, NUMX, lambda j: (f"={P('ebit',j)}/-{P('interest_expense',j)}" if P('ebit',j) and P('interest_expense',j) else None)),
    ]
    r = 2
    for label, sec, fmt, fn in rows:
        if fn is None:                                # section band
            _band(ws, r, label, last_col=last_col, font=F_HEAD, fill_c=MUTE, h=14)
            r += 1
            continue
        _w(ws, f"A{r}", label, font=F_LBL, align="left")
        for j in range(2, ncol + 1):
            f = fn(j)
            _w(ws, f"{get_column_letter(j)}{r}",
               (f"=IFERROR({f[1:]},\"\")" if f else None),
               font=F_VAL, fmt=fmt, align="right", border=True)
        r += 1
    _w(ws, f"A{r + 1}",
       "Every ratio is a live formula off the P&L and Balance Sheet tabs — edit a "
       "statement figure and the ratios move with it.", font=F_NOTE, align="left")


def _xlsx_response(wb: Workbook, filename: str) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    return Response(
        content=data,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"',
                 "Content-Length": str(len(data))},
    )


@router.get("/{ticker}.xlsx")
def export_company(ticker: str, db: Session = Depends(get_db)):
    co = db.query(models.Company).filter_by(ticker=ticker.upper()).first()
    if not co:
        raise HTTPException(404, f"Unknown ticker {ticker}")

    price = co.market.price if co.market else None
    rec, assumptions, analyst, cdata = None, {}, None, None
    try:
        cdata = build_company(db, co)
        assumptions = effective_assumptions(db, co, cdata)
        rec = engines.recommend(cdata, assumptions)
    except Exception:
        rec, assumptions, cdata = None, assumptions or {}, cdata
    try:
        ins = db.query(models.CompanyInsight).filter_by(company_id=co.id).first()
        analyst = analyst_consensus(ins.data if ins else None, price)
    except Exception:
        analyst = None

    hist_rows = (db.query(models.HistoricalFinancial)
                   .filter_by(company_id=co.id).all())
    statements = _shape_statements(hist_rows)

    v = (rec or {}).get("valuation") or {}
    is_fin = (cdata or {}).get("type") == "financial"

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    # Build the live model sheet first so Summary can link to its intrinsic cell.
    # Fallback when no model is built: a plain blended value (never a self-
    # reference into the Summary sheet, which would be circular).
    fair_value = _num((rec or {}).get("intrinsic"), 2)
    model_ws = None
    if v.get("intrinsic") is not None and cdata:
        try:
            model_ws = wb.create_sheet("DCF Model" if not is_fin else "RI Model")
            if is_fin:
                iv_ref = _ri_model_sheet(model_ws, co.ticker, cdata, assumptions, v, price)
            else:
                mature_roic = SP.params(assumptions.get("_valuation_sector")).get("mature_roic") or 0.12
                iv_ref = _fcff_model_sheet(model_ws, co.ticker, cdata, assumptions, v, price, mature_roic)
            fair_value = f"='{model_ws.title}'!{iv_ref}"
        except Exception:
            if model_ws is not None:
                wb.remove(model_ws)

    _summary_sheet(summary, co, price, rec, analyst, fair_value)

    stmt_rows = {}
    for stmt_type, title in (("PL", "P&L"), ("BS", "Balance Sheet"), ("CF", "Cash Flow")):
        stmt_rows[stmt_type] = _statement_sheet(
            wb.create_sheet(title), statements, stmt_type, add_derived=(stmt_type == "PL")) or {}

    # Ratios sheet — live cross-sheet formulas off P&L + Balance Sheet.
    if stmt_rows.get("PL") and stmt_rows.get("BS"):
        _ratios_sheet(wb.create_sheet("Ratios"), sorted(statements.keys()),
                      stmt_rows["PL"], stmt_rows["BS"], "P&L", "Balance Sheet")

    return _xlsx_response(wb, f"{co.ticker}.xlsx")
